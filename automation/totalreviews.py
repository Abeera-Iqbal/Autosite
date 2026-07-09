"""
totalreviews.py
---------------
1. Downloads the latest raw XLSX (IRCR or DCR) from the API.
2. Builds a per-restaurant category breakdown with star counts + complaint ratios.
3. Formats the sheet and saves as BK_Category_Report.xlsx.
4. Uploads BK_Category_Report.xlsx back to the API.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from config import DOWNLOAD_DIR
from drive_helper import upload_file, download_latest_xlsx_files

ARTIFACT_DIR = os.path.abspath("artifacts")
os.makedirs(ARTIFACT_DIR, exist_ok=True)

OUTPUT_FILE = os.path.join(ARTIFACT_DIR, "BK_Category_Report.xlsx")

# ── Read input from downloaded artifact ──────────────────────────────────────
latest_file = os.path.join(ARTIFACT_DIR, "1_2_3_Reviews.xlsx")

if not os.path.exists(latest_file):
    raise FileNotFoundError(
        f"Input artifact not found: {latest_file}"
    )

print("Using artifact:", latest_file)

# ── Read ──────────────────────────────────────────────────────────────────────
df = pd.read_excel(latest_file)
df.columns = df.columns.str.strip()

print("Columns:", df.columns.tolist())

RESTAURANT_COL = "Restaurant #"
MANAGER_COL    = "Area Manager"
CATEGORY_COL   = "Category"
RATING_COL     = "Rating"

CATEGORIES = [
    "Brand Expectation",
    "Customer Service",
    "Food",
    "Operations",
    "Order Compliance",
    "Facility Cleanliness",
    "Just Rating",
    "Delivery Experience",
    "Price & Value",
]

# ── Validate required columns ─────────────────────────────────────────────────
required = [RESTAURANT_COL, MANAGER_COL, CATEGORY_COL, RATING_COL]
missing  = [c for c in required if c not in df.columns]
if missing:
    raise Exception(f"Missing columns: {missing}")

df[RATING_COL] = pd.to_numeric(df[RATING_COL], errors="coerce")

# ── Build result DataFrame ────────────────────────────────────────────────────
unique_restaurants = df[[RESTAURANT_COL, MANAGER_COL]].drop_duplicates()

result_df = pd.DataFrame({
    RESTAURANT_COL: unique_restaurants[RESTAURANT_COL].values,
    MANAGER_COL:    unique_restaurants[MANAGER_COL].values,
})

for star, label in [(1, "1 Star"), (2, "2 Star"), (3, "3 Star")]:
    counts = df[df[RATING_COL] == star].groupby(RESTAURANT_COL).size()
    result_df[label] = result_df[RESTAURANT_COL].map(counts).fillna(0).astype(int)

for cat in CATEGORIES:
    counts = df[df[CATEGORY_COL] == cat].groupby(RESTAURANT_COL).size()
    result_df[cat] = result_df[RESTAURANT_COL].map(counts).fillna(0).astype(int)

result_df["Total Rating"] = (
    result_df["1 Star"] + result_df["2 Star"] + result_df["3 Star"]
)

# ── Summary row ───────────────────────────────────────────────────────────────
month_label = datetime.now().strftime("%b %Y") + " - In Progress"

summary = {RESTAURANT_COL: month_label, MANAGER_COL: ""}
for col in ["1 Star", "2 Star", "3 Star"] + CATEGORIES + ["Total Rating"]:
    summary[col] = result_df[col].sum()

# ── Complaint ratio row ───────────────────────────────────────────────────────
ratio = {RESTAURANT_COL: "Complaints Ratio %", MANAGER_COL: ""}
total_rating = summary["Total Rating"]

for col in ["1 Star", "2 Star", "3 Star"] + CATEGORIES:
    if total_rating == 0:
        ratio[col] = "0%"
    else:
        ratio[col] = f"{round((summary[col] / total_rating) * 100)}%"

ratio["Total Rating"] = ""

# ── Combine ───────────────────────────────────────────────────────────────────
blank = pd.DataFrame([[""] * len(result_df.columns)], columns=result_df.columns)

final_df = pd.concat(
    [result_df, blank, pd.DataFrame([summary]), blank, pd.DataFrame([ratio])],
    ignore_index=True
)

# ── Save raw ──────────────────────────────────────────────────────────────────
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    final_df.to_excel(writer, index=False, sheet_name="BK Category")

# ── Format ────────────────────────────────────────────────────────────────────
wb = load_workbook(OUTPUT_FILE)
ws = wb["BK Category"]

header_fill  = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
summary_fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")

for cell in ws[1]:
    cell.font      = Font(bold=True)
    cell.fill      = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

for column in ws.columns:
    max_len = 0
    col_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_len:
                max_len = len(str(cell.value))
        except Exception:
            pass
    ws.column_dimensions[col_letter].width = max_len + 5

for row in ws.iter_rows():
    first = str(row[0].value)
    if "Progress" in first or "Ratio" in first:
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = summary_fill

wb.save(OUTPUT_FILE)
print("BK_Category_Report.xlsx saved:", OUTPUT_FILE)

# ── Upload to API ─────────────────────────────────────────────────────────────
print("BK_Category_Report.xlsx created successfully.")
print("Output:", OUTPUT_FILE)
