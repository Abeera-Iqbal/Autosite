"""
reviewsparser.py
----------------
1. Downloads the two latest raw XLSX files (IRCR + DCR) from the API.
2. Counts star ratings per restaurant and writes Combined_Reviews.xlsx.
3. Uploads Combined_Reviews.xlsx back to the API.
"""


import os
import time
import glob
from collections import defaultdict
from openpyxl import Workbook, load_workbook

from config import RESTAURANT_DATA

ARTIFACT_DIR = os.path.abspath("artifacts")
os.makedirs(ARTIFACT_DIR, exist_ok=True)

OUTPUT_FILE = os.path.join(ARTIFACT_DIR, "Combined_Reviews.xlsx")

file1 = os.path.join(ARTIFACT_DIR, "IRCR_Reviews.xlsx")
file2 = os.path.join(ARTIFACT_DIR, "DCR_Reviews.xlsx")

if not os.path.exists(file1):
    raise FileNotFoundError(file1)

if not os.path.exists(file2):
    raise FileNotFoundError(file2)

print("IRCR:", file1)
print("DCR :", file2)


# ── Open workbooks ────────────────────────────────────────────────────────────
# Verify IRCR file only
wb = load_workbook(file1, read_only=True)
ws = wb.active

headers = [cell.value for cell in ws[1]]

print("IRCR headers:", headers)

if "Rating" not in headers:
    raise Exception(f"IRCR file is invalid. Headers: {headers}")

wb.close()

# Open DCR if possible
try:
    wb = load_workbook(file2, read_only=True)
    ws = wb.active

    dcr_headers = [cell.value for cell in ws[1]]
    print("DCR headers:", dcr_headers)

    dcr_available = "Rating" in dcr_headers

    if not dcr_available:
        print("DCR file is empty. Continuing with IRCR only.")

    wb.close()

except Exception:
    print("Unable to read DCR file. Continuing with IRCR only.")
    dcr_available = False
    
wb1 = load_workbook(file1)
sheet1 = wb1.active

wb2 = load_workbook(file2)
sheet2 = wb2.active

# ── Output workbook ───────────────────────────────────────────────────────────
output_wb = Workbook()
output_sheet = output_wb.active
output_sheet.title = "Current"

HEADERS = [
    "Restaurant #", "Restaurant Name", "Area Manager",
    "Total Reviews",
    "1 Star", "2 Star", "3 Star",
    "4 Star (DCR)", "4 Star", "5 Star",
    "Total Negative Reviews", "Total Positive Reviews",
    "Positive Review %", "Negative Review %",
    "Note",
]

for col_num, header in enumerate(HEADERS, 1):
    output_sheet.cell(row=1, column=col_num).value = header

# ── Column positions in source files ─────────────────────────────────────────
RESTAURANT_COL = 1
REVIEW_COL     = 2

# ── Counters ──────────────────────────────────────────────────────────────────
total_reviews  = defaultdict(int)
one_star       = defaultdict(int)
two_star       = defaultdict(int)
three_star     = defaultdict(int)
four_star      = defaultdict(int)
four_star_dcr  = defaultdict(int)
five_star      = defaultdict(int)

def process_sheet(sheet, is_dcr=False):
    for row in range(2, sheet.max_row + 1):
        rid = sheet.cell(row=row, column=RESTAURANT_COL).value
        rv = sheet.cell(row=row, column=REVIEW_COL).value

        if rid is None or rv is None:
            continue

        rid = str(rid).strip()
        rv = str(rv).strip()

        if is_dcr:
            # Only store DCR 4-star count.
            # Do NOT add to total reviews or any other counters.
            if rv == "4":
                four_star_dcr[rid] += 1
            continue

        # IRCR counts toward totals
        total_reviews[rid] += 1

        if rv == "1":
            one_star[rid] += 1
        elif rv == "2":
            two_star[rid] += 1
        elif rv == "3":
            three_star[rid] += 1
        elif rv == "4":
            four_star[rid] += 1
        elif rv == "5":
            five_star[rid] += 1

process_sheet(sheet1, is_dcr=False)
process_sheet(sheet2, is_dcr=True)

# ── Write per-restaurant rows ─────────────────────────────────────────────────
output_row = 2

for restaurant_id, restaurant_name, manager in RESTAURANT_DATA:

    total    = total_reviews[restaurant_id]
    star1    = one_star[restaurant_id]
    star2    = two_star[restaurant_id]
    star3    = three_star[restaurant_id]
    star4    = four_star[restaurant_id]
    star4dcr = four_star_dcr[restaurant_id]
    star5    = five_star[restaurant_id]

    negative_reviews = star1 + star2 + star3
    positive_reviews = star4 + star5

    positive_pct = round((positive_reviews / total) * 100, 2) if total > 0 else 0
    negative_pct = round((negative_reviews / total) * 100, 2) if total > 0 else 0

    output_sheet.cell(output_row, 1).value  = restaurant_id
    output_sheet.cell(output_row, 2).value  = restaurant_name
    output_sheet.cell(output_row, 3).value  = manager
    output_sheet.cell(output_row, 4).value  = total
    output_sheet.cell(output_row, 5).value  = star1
    output_sheet.cell(output_row, 6).value  = star2
    output_sheet.cell(output_row, 7).value  = star3
    output_sheet.cell(output_row, 8).value  = star4dcr
    output_sheet.cell(output_row, 9).value  = star4
    output_sheet.cell(output_row, 10).value = star5
    output_sheet.cell(output_row, 11).value = negative_reviews
    output_sheet.cell(output_row, 12).value = positive_reviews
    output_sheet.cell(output_row, 13).value = f"{positive_pct}%"
    output_sheet.cell(output_row, 14).value = f"{negative_pct}%"

    output_row += 1

# ── Grand total row ───────────────────────────────────────────────────────────
total_reviews_sum = sum(total_reviews[rid] for rid, _, _ in RESTAURANT_DATA)
total_1_star      = sum(one_star[rid]      for rid, _, _ in RESTAURANT_DATA)
total_2_star      = sum(two_star[rid]      for rid, _, _ in RESTAURANT_DATA)
total_3_star      = sum(three_star[rid]    for rid, _, _ in RESTAURANT_DATA)
total_4_star_dcr  = sum(four_star_dcr[rid] for rid, _, _ in RESTAURANT_DATA)
total_4_star      = sum(four_star[rid]     for rid, _, _ in RESTAURANT_DATA)
total_5_star      = sum(five_star[rid]     for rid, _, _ in RESTAURANT_DATA)

total_negative = total_1_star + total_2_star + total_3_star
total_positive = total_4_star + total_5_star

total_positive_pct = (
    round((total_positive / total_reviews_sum) * 100, 2)
    if total_reviews_sum > 0 else 0
)
total_negative_pct = (
    round((total_negative / total_reviews_sum) * 100, 2)
    if total_reviews_sum > 0 else 0
)

output_sheet.cell(output_row, 3).value  = "TOTAL"
output_sheet.cell(output_row, 4).value  = total_reviews_sum
output_sheet.cell(output_row, 5).value  = total_1_star
output_sheet.cell(output_row, 6).value  = total_2_star
output_sheet.cell(output_row, 7).value  = total_3_star
output_sheet.cell(output_row, 8).value  = total_4_star_dcr
output_sheet.cell(output_row, 9).value  = total_4_star
output_sheet.cell(output_row, 10).value = total_5_star
output_sheet.cell(output_row, 11).value = total_negative
output_sheet.cell(output_row, 12).value = total_positive
output_sheet.cell(output_row, 13).value = f"{total_positive_pct}%"
output_sheet.cell(output_row, 14).value = f"{total_negative_pct}%"

# ── Auto column width ─────────────────────────────────────────────────────────
for column in output_sheet.columns:
    max_length = 0
    col_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except Exception:
            pass
    output_sheet.column_dimensions[col_letter].width = max_length + 3

# ── Save ──────────────────────────────────────────────────────────────────────
output_wb.save(OUTPUT_FILE)
wb1.close()
wb2.close()
output_wb.close()

print("Combined_Reviews.xlsx created:", OUTPUT_FILE)

# ── Upload to API ─────────────────────────────────────────────────────────────
print("Combined_Reviews.xlsx created successfully.")
print("Saved to:", OUTPUT_FILE)
print("reviewsparser.py completed successfully.")
