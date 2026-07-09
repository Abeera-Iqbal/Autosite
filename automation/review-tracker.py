import os
import glob
from openpyxl import Workbook, load_workbook

# Folder where dennys.py saves the downloaded Excel
ARTIFACT_DIR = os.path.abspath("artifacts")
os.makedirs(ARTIFACT_DIR, exist_ok=True)

print("Artifact Directory:", ARTIFACT_DIR)

print("Files inside artifacts:")
for f in os.listdir(ARTIFACT_DIR):
    print(" -", f)

excel_files = glob.glob(os.path.join(ARTIFACT_DIR, "Review Feed*.xlsx"))

OUTPUT_FILE = os.path.join(
    ARTIFACT_DIR,
    "Review_Tracker.xlsx"
)

if not excel_files:
    raise FileNotFoundError(
        f"No Review Feed Excel file found in {ARTIFACT_DIR}"
    )

excel_files.sort(key=os.path.getmtime, reverse=True)

input_file = excel_files[0]

print("Reading:", input_file)

if not excel_files:
    raise Exception("No Excel files found.")

excel_files.sort(key=os.path.getctime, reverse=True)
input_file = excel_files[0]

print(f"Reading: {input_file}")

# OPEN SOURCE FILE

source_wb = load_workbook(input_file, data_only=True)
source_sheet = source_wb.active

print("Reading:", input_file)

# CREATE / OPEN TRACKER FILE

if os.path.exists(OUTPUT_FILE):

    output_wb = load_workbook(OUTPUT_FILE)

    positive_sheet = output_wb["Positive Reviews"]
    negative_sheet = output_wb["Negative Reviews"]

else:

    output_wb = Workbook()

    positive_sheet = output_wb.active
    positive_sheet.title = "Positive Reviews"

    negative_sheet = output_wb.create_sheet("Negative Reviews")

    positive_headers = [
        "Restaurant",
        "Address",
        "City",
        "Date",
        "Star Rating",
        "Review",
        "Author",
        "Owner Response Date",
        "Owner Response"
    ]

    negative_headers = [
        "Restaurant",
        "Rating",
        "Review",
        "Review Date",
        "Category",
        "Shift",
        "Action Plan",
        "Closure",
        "Owner Response Date",
        "Owner Response"
    ]

    for col, header in enumerate(positive_headers, start=1):
        positive_sheet.cell(1, col).value = header

    for col, header in enumerate(negative_headers, start=1):
        negative_sheet.cell(1, col).value = header

# EXISTING REVIEWS


existing_positive = set()
existing_negative = set()

for row in range(2, positive_sheet.max_row + 1):

    review = positive_sheet.cell(row=row, column=6).value

    if review:
        existing_positive.add(str(review).strip())

for row in range(2, negative_sheet.max_row + 1):

    review = negative_sheet.cell(row=row, column=3).value

    if review:
        existing_negative.add(str(review).strip())

# COLLECT REVIEWS


positive_rows = []
negative_rows = []

for row in range(2, source_sheet.max_row + 1):

    address = source_sheet.cell(row=row, column=6).value
    city = source_sheet.cell(row=row, column=7).value
    review_date = source_sheet.cell(row=row, column=11).value
    review = source_sheet.cell(row=row, column=12).value
    author = source_sheet.cell(row=row, column=13).value
    rating = source_sheet.cell(row=row, column=14).value
    owner_response_date = source_sheet.cell(row=row, column=16).value
    owner_response = source_sheet.cell(row=row, column=17).value

    if rating is None:
        continue

    try:
        rating = int(float(rating))
    except:
        continue

    review = str(review).strip() if review else ""

    if review == "":
        continue

    # DETERMINE RESTAURANT


    restaurant = ""

    if address:

        addr = str(address).lower()

        if "jarrell" in addr:
            restaurant = "Jarrell"

        elif "hutto" in addr:
            restaurant = "Hutto"

        elif "bastrop" in addr:
            restaurant = "Bastrop"

        elif "austin" in addr:
            restaurant = "Techridge-Austin"

        else:
            restaurant = str(address)

    # NEGATIVE REVIEWS (1-3)

    if rating <= 3:

        if review in existing_negative:
            continue

        negative_rows.append([
            restaurant,
            rating,
            review,
            review_date,
            "",
            "",
            "",
            "",
            owner_response_date,
            owner_response
        ])

    # POSITIVE REVIEWS (4-5)

    else:

        if review in existing_positive:
            continue

        positive_rows.append([
            restaurant,
            address,
            city,
            review_date,
            rating,
            review,
            author,
            owner_response_date,
            owner_response
        ])

# ==================================================
# INSERT POSITIVE AT TOP
# ==================================================

if positive_rows:

    positive_sheet.insert_rows(2, len(positive_rows))

    for r, row_data in enumerate(positive_rows, start=2):

        for c, value in enumerate(row_data, start=1):

            positive_sheet.cell(r, c).value = value

# INSERT NEGATIVE AT TOP

if negative_rows:

    negative_sheet.insert_rows(2, len(negative_rows))

    for r, row_data in enumerate(negative_rows, start=2):

        for c, value in enumerate(row_data, start=1):

            negative_sheet.cell(r, c).value = value

# AUTO WIDTH

for sheet in [positive_sheet, negative_sheet]:

    for column in sheet.columns:

        max_length = 0

        for cell in column:

            if cell.value is not None:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        sheet.column_dimensions[
            column[0].column_letter
        ].width = min(max_length + 3, 80)

# SAVE

output_wb.save(OUTPUT_FILE)

source_wb.close()
output_wb.close()

print()
print(f"Tracker Updated: {OUTPUT_FILE}")
print(f"New Positive Reviews: {len(positive_rows)}")
print(f"New Negative Reviews: {len(negative_rows)}")
