import os
import sys
from collections import defaultdict
import glob
from openpyxl import Workbook, load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import RUN_ID, TARGET_YEAR, TARGET_MONTH
from drive_helper import upload_file
from json_export import xlsx_to_json

ARTIFACT_DIR = os.path.abspath("artifacts")
os.makedirs(ARTIFACT_DIR, exist_ok=True)

SOURCE_FILE = os.path.join(
    ARTIFACT_DIR,
    "Review_Tracker.xlsx"
)

OUTPUT_FILE = os.path.join(
    ARTIFACT_DIR,
    "Dennys_record.xlsx"
)

PERIOD_TAG = f"{TARGET_YEAR}-{TARGET_MONTH:02d}"
UPLOAD_BASENAME = f"Dennys_record_{PERIOD_TAG}_{RUN_ID}"

if not os.path.exists(SOURCE_FILE):
    raise FileNotFoundError(
        f"{SOURCE_FILE} not found."
    )

print("Reading:", SOURCE_FILE)

wb = load_workbook(
    SOURCE_FILE,
    data_only=True
)

positive_sheet = wb["Positive Reviews"]
negative_sheet = wb["Negative Reviews"]

# SUMMARY

summary = defaultdict(
    lambda: {
        "negative": 0,
        "positive": 0,
        "total": 0
    }
)

# POSITIVE REVIEWS

for row in range(2, positive_sheet.max_row + 1):

    restaurant = positive_sheet.cell(row, 1).value
    review_date = positive_sheet.cell(row, 4).value
    rating = positive_sheet.cell(row, 5).value

    if not restaurant or rating is None:
        continue

    try:
        rating = int(float(rating))
    except:
        continue

    key = (str(restaurant).strip(), str(review_date))

    summary[key]["total"] += 1
    summary[key]["positive"] += 1

# NEGATIVE REVIEWS

for row in range(2, negative_sheet.max_row + 1):

    restaurant = negative_sheet.cell(row, 1).value
    review_date = negative_sheet.cell(row, 4).value
    rating = negative_sheet.cell(row, 2).value

    if not restaurant or rating is None:
        continue

    try:
        rating = int(float(rating))
    except:
        continue

    key = (str(restaurant).strip(), str(review_date))

    summary[key]["total"] += 1
    summary[key]["negative"] += 1


# CREATE OUTPUT

output_wb = Workbook()
output_sheet = output_wb.active
output_sheet.title = "Current"

headers = [
    "Restaurant",
    "Date",
    "Negative Count",
    "Positive Count",
    "Total Reviews"
]

for col, header in enumerate(headers, start=1):
    output_sheet.cell(1, col).value = header


# WRITE DATA

output_row = 2
current_date = None

sorted_data = sorted(
    summary.items(),
    key=lambda x: str(x[0][1]),
    reverse=True
)

for (restaurant, review_date), data in sorted_data:

    if current_date is not None and review_date != current_date:
        output_row += 1

    current_date = review_date

    output_sheet.cell(output_row, 1).value = restaurant
    output_sheet.cell(output_row, 2).value = review_date
    output_sheet.cell(output_row, 3).value = data["negative"]
    output_sheet.cell(output_row, 4).value = data["positive"]
    output_sheet.cell(output_row, 5).value = data["total"]

    output_row += 1

# AUTO WIDTH


for column in output_sheet.columns:

    max_length = 0

    for cell in column:

        if cell.value is not None:
            max_length = max(
                max_length,
                len(str(cell.value))
            )

    output_sheet.column_dimensions[
        column[0].column_letter
    ].width = max_length + 3

# SAVE

output_wb.save(OUTPUT_FILE)

wb.close()
output_wb.close()

print(f"Report Created: {OUTPUT_FILE}")

# ── JSON export + upload (tagged with period + run id) ────────────────────────
output_json = xlsx_to_json(OUTPUT_FILE)

upload_file(OUTPUT_FILE, upload_name=f"{UPLOAD_BASENAME}.xlsx")
upload_file(output_json, upload_name=f"{UPLOAD_BASENAME}.json")

print("Dennys_record uploaded:", UPLOAD_BASENAME)
