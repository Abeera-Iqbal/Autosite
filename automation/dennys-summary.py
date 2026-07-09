import os
import sys
from collections import defaultdict
from openpyxl import Workbook, load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import RUN_ID, TARGET_YEAR, TARGET_MONTH
from drive_helper import upload_file
from json_export import xlsx_to_json

ARTIFACT_DIR = os.path.abspath("artifacts")
os.makedirs(ARTIFACT_DIR, exist_ok=True)

SOURCE_FILE = os.path.join(
    ARTIFACT_DIR,
    "Review_Tracker.xlsx"
)

OUTPUT_FILE = os.path.join(
    ARTIFACT_DIR,
    "Dennys_Summary.xlsx"
)

PERIOD_TAG = f"{TARGET_YEAR}-{TARGET_MONTH:02d}"
UPLOAD_BASENAME = f"Dennys_Summary_{PERIOD_TAG}_{RUN_ID}"

if not os.path.exists(SOURCE_FILE):
    raise FileNotFoundError(
        f"{SOURCE_FILE} not found."
    )

print("Reading:", SOURCE_FILE)

# OPEN TRACKER FILE

wb = load_workbook(SOURCE_FILE, data_only=True)

positive_sheet = wb["Positive Reviews"]
negative_sheet = wb["Negative Reviews"]

# STORAGE


one_star = defaultdict(int)
two_star = defaultdict(int)
three_star = defaultdict(int)
four_star = defaultdict(int)
five_star = defaultdict(int)

total_reviews = defaultdict(int)

# READ POSITIVE REVIEWS SHEET
# Rating = Column E
# Restaurant = Column A

for row in range(2, positive_sheet.max_row + 1):

    location = positive_sheet.cell(row=row, column=1).value
    rating = positive_sheet.cell(row=row, column=5).value

    if location is None or rating is None:
        continue

    location = str(location).strip()

    try:
        rating = int(float(rating))
    except:
        continue

    total_reviews[location] += 1

    if rating == 4:
        four_star[location] += 1

    elif rating == 5:
        five_star[location] += 1

# READ NEGATIVE REVIEWS SHEET
# Rating = Column B
# Restaurant = Column A

for row in range(2, negative_sheet.max_row + 1):

    location = negative_sheet.cell(row=row, column=1).value
    rating = negative_sheet.cell(row=row, column=2).value

    if location is None or rating is None:
        continue

    location = str(location).strip()

    try:
        rating = int(float(rating))
    except:
        continue

    total_reviews[location] += 1

    if rating == 1:
        one_star[location] += 1

    elif rating == 2:
        two_star[location] += 1

    elif rating == 3:
        three_star[location] += 1

# CREATE OUTPUT FILE

output_wb = Workbook()
output_sheet = output_wb.active
output_sheet.title = "Summary"

headers = [
    "Location",
    "1 Star",
    "2 Star",
    "3 Star",
    "4 Star",
    "5 Star",
    "Total Reviews",
    "Negative Reviews",
    "Positive Reviews"
]

for col, header in enumerate(headers, start=1):
    output_sheet.cell(row=1, column=col).value = header

# WRITE DATA

row_num = 2

for location in sorted(total_reviews.keys()):

    star1 = one_star[location]
    star2 = two_star[location]
    star3 = three_star[location]
    star4 = four_star[location]
    star5 = five_star[location]

    total = total_reviews[location]

    negative = star1 + star2 + star3
    positive = star4 + star5

    output_sheet.cell(row=row_num, column=1).value = location
    output_sheet.cell(row=row_num, column=2).value = star1
    output_sheet.cell(row=row_num, column=3).value = star2
    output_sheet.cell(row=row_num, column=4).value = star3
    output_sheet.cell(row=row_num, column=5).value = star4
    output_sheet.cell(row=row_num, column=6).value = star5
    output_sheet.cell(row=row_num, column=7).value = total
    output_sheet.cell(row=row_num, column=8).value = negative
    output_sheet.cell(row=row_num, column=9).value = positive

    row_num += 1

# GRAND TOTAL

output_sheet.cell(row=row_num, column=1).value = "TOTAL"

output_sheet.cell(row=row_num, column=2).value = sum(one_star.values())
output_sheet.cell(row=row_num, column=3).value = sum(two_star.values())
output_sheet.cell(row=row_num, column=4).value = sum(three_star.values())
output_sheet.cell(row=row_num, column=5).value = sum(four_star.values())
output_sheet.cell(row=row_num, column=6).value = sum(five_star.values())

output_sheet.cell(row=row_num, column=7).value = sum(total_reviews.values())

output_sheet.cell(row=row_num, column=8).value = (
    sum(one_star.values())
    + sum(two_star.values())
    + sum(three_star.values())
)

output_sheet.cell(row=row_num, column=9).value = (
    sum(four_star.values())
    + sum(five_star.values())
)

# AUTO WIDTH
for column in output_sheet.columns:

    max_length = 0
    column_letter = column[0].column_letter

    for cell in column:

        if cell.value is not None:
            max_length = max(
                max_length,
                len(str(cell.value))
            )

    output_sheet.column_dimensions[column_letter].width = max_length + 3

# SAVE

output_wb.save(OUTPUT_FILE)

wb.close()
output_wb.close()

print()
print("Summary created successfully")
print("Output:", OUTPUT_FILE)

# ── JSON export + upload (tagged with period + run id) ────────────────────────
output_json = xlsx_to_json(OUTPUT_FILE)

upload_file(OUTPUT_FILE, upload_name=f"{UPLOAD_BASENAME}.xlsx")
upload_file(output_json, upload_name=f"{UPLOAD_BASENAME}.json")

print("Dennys_Summary uploaded:", UPLOAD_BASENAME)
